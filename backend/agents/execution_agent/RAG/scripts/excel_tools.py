from __future__ import annotations

import glob
import os
from datetime import datetime
from typing import Sequence

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def _get_agent_folder(subfolder: str) -> str:
    possible_paths = [
        os.path.expanduser("~/OneDrive/Desktop/agent"),
        os.path.expanduser("~/Desktop/agent"),
        os.path.expanduser("~/Documents/agent"),
    ]

    for base_path in possible_paths:
        if os.path.exists(os.path.dirname(base_path)) or os.path.exists(base_path):
            folder = os.path.join(base_path, subfolder)
            os.makedirs(folder, exist_ok=True)
            return folder

    fallback_path = os.path.join(os.path.expanduser("~"), "agent", subfolder)
    os.makedirs(fallback_path, exist_ok=True)
    return fallback_path


def _ensure_xlsx_extension(name: str) -> str:
    if name.lower().endswith(".xlsx"):
        return name
    return f"{name}.xlsx"


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _print_success(path: str) -> None:
    print(f"[FILE]: {path}")
    print("EXECUTION_SUCCESS")


def _save_with_retry(wb: Workbook, path: str) -> str:
    try:
        wb.save(path)
        return path
    except PermissionError:
        base, ext = os.path.splitext(path)
        retry_path = f"{base}_v2{ext or '.xlsx'}"
        wb.save(retry_path)
        return retry_path


def _resolve_excel_path(filename: str) -> str:
    if os.path.isabs(filename) and os.path.exists(filename):
        return filename

    folder = _get_agent_folder("excel")
    search_roots = [
        os.path.expanduser("~\\Desktop"),
        os.path.expanduser("~\\OneDrive\\Desktop"),
        os.path.expanduser("~\\Documents"),
        os.path.expanduser("~\\Downloads"),
        folder,
    ]

    for root in search_roots:
        if not os.path.exists(root):
            continue
        direct = os.path.join(root, filename)
        if os.path.isfile(direct):
            return direct
        matches = glob.glob(os.path.join(root, "**", filename), recursive=True)
        if matches:
            return matches[0]

    raise FileNotFoundError(f"Could not find file: {filename}")


def xl_create(name: str | None = None) -> str:
    folder = _get_agent_folder("excel")
    filename = _ensure_xlsx_extension(name) if name else f"spreadsheet_{_timestamp()}.xlsx"
    save_path = filename if os.path.isabs(filename) else os.path.join(folder, filename)

    wb = Workbook()
    final_path = _save_with_retry(wb, save_path)
    _print_success(final_path)
    return final_path


def xl_open(path: str) -> None:
    target = _resolve_excel_path(path)
    os.startfile(target)
    _print_success(target)


def xl_find(filename: str) -> str:
    found = _resolve_excel_path(filename)
    _print_success(found)
    return found


def xl_load(path: str) -> Workbook:
    from openpyxl import load_workbook
    resolved = _resolve_excel_path(path)
    return load_workbook(resolved)


def xl_set_cell(ws: Worksheet, row: int, col: int, value: str | int | float) -> Worksheet:
    ws.cell(row=row, column=col, value=value)
    return ws


def xl_write_headers(ws: Worksheet, headers: Sequence[str], row: int = 1) -> Worksheet:
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=str(header))
    return ws


def xl_write_row(ws: Worksheet, row: int, values: Sequence[str | int | float]) -> Worksheet:
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=row, column=col_idx, value=value)
    return ws


def xl_set_formula(ws: Worksheet, row: int, col: int, formula: str) -> Worksheet:
    ws.cell(row=row, column=col, value=formula)
    return ws


def xl_save(wb: Workbook, path: str) -> str:
    save_path = _ensure_xlsx_extension(path)
    if not os.path.isabs(save_path):
        save_path = os.path.join(_get_agent_folder("excel"), save_path)
    final_path = _save_with_retry(wb, save_path)
    _print_success(final_path)
    return final_path


def xl_launch() -> None:
    os.system("start excel")
    print("EXECUTION_SUCCESS")


__all__ = [
    "xl_create",
    "xl_open",
    "xl_find",
    "xl_load",
    "xl_set_cell",
    "xl_write_headers",
    "xl_write_row",
    "xl_set_formula",
    "xl_save",
    "xl_launch",
]
